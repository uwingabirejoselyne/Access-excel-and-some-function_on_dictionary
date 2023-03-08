import openpyxl
wb = openpyxl.load_workbook('risks.xlsx')
sheet=wb["Sheet1"]
# data=[]
# for row_data in sheet.iter_rows():
#     values={}
#     if row_data[0].value==country:
#         values['Floods']=row_data[1].value
#         values['Wild Fires']=row_data[2].value
#         values['Storms']=row_data[3].value
#         values['Droughts']=row_data[4].value
#         values['Rise in sea level']=row_data[5].value
#         values['Percipitation']=row_data[6].value
#         values['Landslides']=row_data[7].value
#         values['Heat Waves']=row_data[8].value
#         data.append(values)

def getCountry(country):
    # val=[]
    # for i in data:
    #     if i["Country"]==country:
    #         val=i
    # return val
    values={}
    for row_data in sheet.iter_rows():
        
        if row_data[0].value==country:
            values['Floods']=row_data[1].value
            values['Wild Fires']=row_data[2].value
            values['Storms']=row_data[3].value
            values['Droughts']=row_data[4].value
            values['Rise in sea level']=row_data[5].value
            values['Percipitation']=row_data[6].value
            values['Landslides']=row_data[7].value
            values['Heat Waves']=row_data[8].value
            
    return values

ABC=getCountry('Rwanda')
print(ABC)
 =list(ABC.keys())
yaxis=list(ABC.values())
print("x-axix: ",xaxis)
print("y-axix: ",yaxis)