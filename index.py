# import pandas as pd
# df = pd.read_excel('music.xlsx')
# df['Total'] = df.iloc[:, 4:9].sum(axis=1)
# df.to_excel('modified.xlsx', index = False)
# print(df)

# def top_three_student(student_lists):
#     top_three = []
#     for i in range(3):
#         max_age = 0
#         max_Student = None
#         for student in student_lists:
#             if student[1]>max_age and student not in top_three:
#                 max_age = student[1]
#                 max_Student = student
#         top_three.append(max_Student)
#     return top_three


# student = [('jose',4),("ali",6),("kelly",7),('titty',8)]
# top_students = top_three_student(student)
# print(top_students)
# student = {
#     "A":20,
#     "b":7,
#     "C": 8
# }
# names = []
# ages = []
# for name,age in student.items():
#     names.append(name)
#     ages.append(age  )
# print(f'names = {names}')
# print(f'ages {ages}')

# task ={}
# for i in range(len(names)):
#     task[names[i]] = ages[i]
# print(f'{task}')

# def top_three_student(student_lists):
#     top_three = []
#     for i in range(3):
#         max_age = 0
#         max_Student = max(student_lists)
#         for student in student_lists:
#             if student[1]>max_age and student not in top_three:
#                 max_age = student[1]
#                 max_Student = student
#         top_three.append(max_Student)
#     return top_three


# student =[('Joselyne',12),('Uwingabire',5),('Keza',9),('Ali',10)]
# top_students =top_three_student(student)
# print("top_three_student is :", top_students)


import pandas as pd
# students = [('Ankit', 22, 'A'),('swapnil', 22, 'A'),('priya', 22, 'A'),('shivangi', 22, 'A'),]

# # create dataframe objects 

# stu_df = pd.DataFrame(students, columns = ['name','age','section'],index = ['1','2','3','4'])
# print(stu_df)


# data = {'Name':['Jai', 'Princi', 'Gaurav', 'Anuj'],
# 'Age':[27, 24, 22,32],
# 'Addresss':['Delhi', 'Kannpur','Allahabad','Kannauj'],
# 'Qualification':['Msc', 'Ma','Mca','phd'],
# }
# df = pd.DataFrame(data)

# location = ['Kigali','Nairobi','Kampala','Bujumbura']
# df['Address '] = location
# df.drop(['Age','Addresss'], axis =1, inplace = True)

# print (df)


# df = pd.read_excel('music.xlsx')
# for col in df:
#     # print("column", col)
#     print("first 3 row", df[col][:4])

data = pd.read_excel('employees.xlsx')
# name=list(data.iloc[:,0])
# gender = list(data.iloc[:,1])
# print("firstname =",name,"gender =",gender)

# value ={}
# for i in range(0,len(fname)):
#     value[fname[i]]=gender[i]
# for fname,gender in value.items():
#     print("FirstName = ",fname ,"gender =",gender)
#print(data)

# print(data.loc[data['First Name']=='Lisa'])

def Accessdata(name):
    data = pd.read_excel('employees.xlsx')
    d=data.loc[data['First Name']==name]
    return d

abc=Accessdata("Lisa")
#print(dict(abc))
#key=list(dict(abc).values())
name=abc["Gender"]
# dic={}
# for i in abc:
#     list=abc[i])
import openpyxl
wb = openpyxl.load_workbook('employees.xlsx')
sheet=wb["abc"]
data=[]
for row_data in sheet.iter_rows():
    values={}
    values['name']=row_data[0].value
    values['gender']=row_data[1].value
    values['date']=row_data[2].value
    data.append(values)
print(data)


import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active
c1 = sheet.cell(row = 1, column = 1)
c1.value = "Name"
  
c2 = sheet.cell(row= 1 , column = 2)
c2.value = "Age"

c3 = sheet.cell(row= 1 , column = 3)
c3.value = "Country"

c4 = sheet['A2']
c4.value = "RAHUL"

c5 = sheet['B2']
c5.value = "10"

c6= sheet['C2']
c6.value = "Rwanda"

c7 = sheet['A3']
c7.value = "Joselyne"

c8 = sheet['B3']
c8.value = "20"

c9= sheet['C3']
c9.value = "USA"

c10 = sheet['A4']
c10.value = "Keza"

c11 = sheet['B4']
c11.value = "40"

c12= sheet['C4']
c12.value = "Tanzania"
wb.save("demo.xlsx")


from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = 'formatting'

income = [('Income', '6000'),
          ('Salary', 1000),
          ('Investment', 500),
          ('Side hustle', 500),
          ('Total', 2000),
          ('Expense','1200'),
           ('Housing', 1200),
           ('Insurance', 200),
           ('Grocery', 500),
           ('Entertainment', 500),
           ('Total', 2400)]

for row in income:
    ws.append(row)

wb.save('formatting.xlsx')